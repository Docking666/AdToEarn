"""
广告账户审计模块 (Audit)
参照 Claude-ads 的设计思路，对广告投放效果进行可视化数据分析：
  - 投放数据总览（健康评分 + 关键指标）
  - 时间维度趋势分析
  - 账户维度数据对比
  - 异常 / 风险提示（分级：critical / high / medium / low）

数据模型（双轨制，Phase1 升级）：
  每条记录 = { raw: {原始所有字段}, mapped: {映射后标准字段}, tags: {} }
  raw   : 用户上传 CSV 的原始字段，全部保留，不做任何强制映射
  mapped: 经字段映射引擎处理后的 7 个标准字段（account/date/impressions/clicks/
          conversions/spend/conversion_value），用于指标计算与异常检测
  tags  : 行级标签（Phase2 批量打标用），结构 {group_name: [tag_values]}

派生指标：CTR / CVR / CPC / CPM / CPA / ROAS

数据来源：
  1) WebUI 上传 CSV（推荐列：account,date,impressions,clicks,conversions,spend,conversion_value）
  2) WebUI 上传 JSON（数组形式，字段同上）
  3) 生成示例数据（sample=true，演示/测试用）
持久化：JSON 文件（settings.audit_data_file），与 api_config.json 同风格
"""

import csv
import io
import json
import random
import statistics
import threading
import time
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Optional

from ..config import settings
from .app_logger import log_collector, EVENT_SYSTEM, EVENT_CONFIG
from .field_mapper import field_mapper

# 严重度级别
SEV_CRITICAL = "critical"
SEV_HIGH = "high"
SEV_MEDIUM = "medium"
SEV_LOW = "low"

SEV_LABELS = {
    SEV_CRITICAL: "严重",
    SEV_HIGH: "高危",
    SEV_MEDIUM: "中等",
    SEV_LOW: "提示",
}

# 指标元信息（供前端展示：名称/单位/精度/方向）
METRIC_META = {
    "impressions": {"label": "曝光量", "unit": "", "precision": 0, "format": "num"},
    "clicks": {"label": "点击量", "unit": "", "precision": 0, "format": "num"},
    "conversions": {"label": "转化量", "unit": "", "precision": 0, "format": "num"},
    "spend": {"label": "花费", "unit": "¥", "precision": 2, "format": "money"},
    "conversion_value": {"label": "转化价值", "unit": "¥", "precision": 2, "format": "money"},
    "ctr": {"label": "点击率 CTR", "unit": "%", "precision": 2, "format": "percent"},
    "cvr": {"label": "转化率 CVR", "unit": "%", "precision": 2, "format": "percent"},
    "cpc": {"label": "单次点击成本 CPC", "unit": "¥", "precision": 2, "format": "money"},
    "cpm": {"label": "千次曝光成本 CPM", "unit": "¥", "precision": 2, "format": "money"},
    "cpa": {"label": "获客成本 CPA", "unit": "¥", "precision": 2, "format": "money"},
    "roas": {"label": "投产比 ROAS", "unit": "", "precision": 2, "format": "ratio"},
}


class AuditService:
    """广告账户审计服务（数据管理 + 指标计算 + 异常检测）"""

    def __init__(self):
        self._lock = threading.Lock()
        self._data_file: Path = settings.audit_data_file
        self._history_dir: Path = settings.audit_data_file.parent / "audit_history"

    # ==================== 数据读写 ====================
    def _load(self) -> list:
        """读取全部投放记录"""
        if self._data_file.exists():
            try:
                data = json.loads(self._data_file.read_text(encoding="utf-8"))
                if isinstance(data, list):
                    return data
            except (json.JSONDecodeError, OSError):
                pass
        return []

    def _save(self, records: list):
        """持久化投放记录"""
        self._data_file.parent.mkdir(parents=True, exist_ok=True)
        self._data_file.write_text(
            json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8"
        )

    # ==================== 数据管理 ====================
    def clear(self) -> dict:
        """清空全部审计数据（清空前自动归档快照）"""
        with self._lock:
            records = self._load()
            if records:
                self._archive_snapshot("clear")
            self._save([])
        log_collector.info(EVENT_CONFIG, "审计数据已清空")
        return {"ok": True, "count": 0}

    # ==================== 快照归档（Phase11：稀疏时序） ====================
    def _archive_snapshot(self, reason: str) -> Optional[str]:
        """归档当前数据为历史快照（导入/打标/清空前自动调用）
        返回快照 id；数据为空或归档失败返回 None
        """
        records = self._load()
        if not records:
            return None
        accounts = sorted({r.get("mapped", {}).get("account", "") for r in records if r.get("mapped", {}).get("account")})
        dates = sorted(r.get("mapped", {}).get("date", "") for r in records if r.get("mapped", {}).get("date"))
        snapshot = {
            "imported_at": datetime.now().isoformat(timespec="seconds"),
            "reason": reason,
            "record_count": len(records),
            "accounts": accounts,
            "date_range": [dates[0], dates[-1]] if dates else [],
            "records": records,
        }
        try:
            sid = datetime.now().strftime("%Y%m%d_%H%M%S")
            self._history_dir.mkdir(parents=True, exist_ok=True)
            (self._history_dir / f"{sid}.json").write_text(
                json.dumps(snapshot, ensure_ascii=False), encoding="utf-8")
            self._prune_snapshots()
            log_collector.info(EVENT_CONFIG, f"审计快照已归档: {sid}（{reason}，{len(records)} 条）")
            return sid
        except OSError:
            return None

    def list_snapshots(self) -> list:
        """列出全部历史快照元信息（按时间倒序）"""
        if not self._history_dir.exists():
            return []
        out = []
        for f in sorted(self._history_dir.glob("*.json"), reverse=True):
            try:
                d = json.loads(f.read_text(encoding="utf-8"))
                out.append({
                    "id": f.stem,
                    "imported_at": d.get("imported_at", ""),
                    "reason": d.get("reason", ""),
                    "record_count": d.get("record_count", 0),
                    "accounts": d.get("accounts", []),
                    "date_range": d.get("date_range", []),
                })
            except (json.JSONDecodeError, OSError):
                continue
        return out

    def load_snapshot(self, sid: str) -> Optional[dict]:
        """读取快照完整内容（含 records）"""
        try:
            f = self._history_dir / f"{sid}.json"
            if not f.exists():
                return None
            return json.loads(f.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return None

    def delete_snapshot(self, sid: str) -> bool:
        try:
            f = self._history_dir / f"{sid}.json"
            if not f.exists():
                return False
            f.unlink()
            log_collector.info(EVENT_CONFIG, f"审计快照已删除: {sid}")
            return True
        except OSError as e:
            log_collector.warn(EVENT_CONFIG, f"删除快照失败({sid}): {e}")
            return False

    def _prune_snapshots(self, keep: int = 30):
        """自动裁剪：只保留最近 keep 份快照，防膨胀"""
        try:
            if not self._history_dir.exists():
                return
            files = sorted(self._history_dir.glob("*.json"), reverse=True)
            for f in files[keep:]:
                f.unlink()
        except OSError:
            pass

    def _load_snapshot_records(self, sid: str) -> list:
        """读取快照中的 records（供 advisor 序列分析用）"""
        snap = self.load_snapshot(sid)
        return snap.get("records", []) if snap else []

    def get_meta(self) -> dict:
        """数据元信息：记录数 / 账户列表 / 时间范围 / 是否含示例数据 / 原始字段列表"""
        records = self._load()
        accounts = sorted({r.get("mapped", {}).get("account", "") for r in records if r.get("mapped", {}).get("account")})
        dates = sorted(r.get("mapped", {}).get("date", "") for r in records if r.get("mapped", {}).get("date"))
        has_sample = any(r.get("sample") for r in records)
        # 收集所有 raw 字段（供前端表格列展示）
        raw_fields = []
        seen = set()
        for r in records:
            for k in r.get("raw", {}).keys():
                if k not in seen:
                    seen.add(k)
                    raw_fields.append(k)
        return {
            "record_count": len(records),
            "accounts": accounts,
            "date_min": dates[0] if dates else None,
            "date_max": dates[-1] if dates else None,
            "has_sample": has_sample,
            "raw_fields": raw_fields,
            "metric_meta": METRIC_META,
            "severity_labels": SEV_LABELS,
            "has_creative": any(r.get("mapped", {}).get("creative") for r in records),  # Phase11
        }

    def import_records(self, records: list, source: str = "upload") -> dict:
        """批量导入记录（覆盖式：导入即替换，保持与 api_config 保存语义一致）
        Phase1 升级：自动调用 field_mapper 将原始列名映射到标准字段
        """
        normalized = []
        errors = []
        # 第 1 行用于推断列名 → 字段映射（同批数据假设列名一致）
        column_map = {}
        if records and isinstance(records[0], dict):
            for col in records[0].keys():
                m = field_mapper.match(col)
                if m["standard_field"]:
                    column_map[col] = m["standard_field"]
        for i, raw in enumerate(records):
            try:
                rec = self._normalize(raw, column_map)
                if rec:
                    normalized.append(rec)
            except ValueError as e:
                errors.append({"row": i + 1, "error": str(e)})
        if not normalized:
            return {"ok": False, "imported": 0, "errors": errors or [{"row": 1, "error": "未解析到有效数据，请检查列名与格式"}]}
        with self._lock:
            self._archive_snapshot("import")  # Phase11: 导入前归档旧数据
            self._save(normalized)
        log_collector.info(EVENT_CONFIG, f"审计数据导入成功: {len(normalized)} 条", {
            "source": source, "errors": len(errors),
            "mapped_fields": list(column_map.keys()),
        })
        return {"ok": True, "imported": len(normalized), "errors": errors, "column_map": column_map}

    def _normalize(self, raw: dict, column_map: Optional[dict] = None) -> Optional[dict]:
        """单条记录标准化与校验
        Phase1 双轨制：保留 raw 原始所有字段 + 构建 mapped 标准字段
        column_map: 预先匹配好的 {原始列名: 标准字段} 映射；为 None 时按列名逐个匹配
        """
        if not isinstance(raw, dict):
            raise ValueError("记录必须是对象")

        # raw: 原始所有字段保留
        raw_clean = {k: v for k, v in raw.items() if v is not None and v != ""}

        # 构建 mapped: 标准字段
        mapped = {}
        for raw_col, value in raw_clean.items():
            if column_map and raw_col in column_map:
                std_field = column_map[raw_col]
            elif column_map is None:
                m = field_mapper.match(raw_col)
                std_field = m["standard_field"]
            else:
                std_field = None
            if std_field and std_field not in mapped:
                mapped[std_field] = value

        # 必需字段校验（account/date）
        if not mapped.get("account"):
            mapped["account"] = "未分组账户"
        if not mapped.get("date"):
            raise ValueError("缺少日期字段（需映射到 date 标准字段）")
        try:
            d = self._parse_date(str(mapped["date"]))
            mapped["date"] = d.isoformat()
        except ValueError:
            raise ValueError(f"日期格式无效: {mapped['date']}（应为 YYYY-MM-DD）")
        mapped["account"] = str(mapped["account"]).strip() or "未分组账户"
        # Phase11: 素材名清洗（可选字段；无则 None）
        if mapped.get("creative"):
            mapped["creative"] = str(mapped["creative"]).strip()
            if not mapped["creative"]:
                del mapped["creative"]

        # 数值字段：清洗 + 类型转换
        for num_field in ("impressions", "clicks", "conversions", "spend", "conversion_value"):
            if num_field in mapped:
                try:
                    v = str(mapped[num_field]).replace(",", "").replace("¥", "").replace("元", "").replace("$", "").strip()
                    mapped[num_field] = float(v) if num_field in ("spend", "conversion_value") else int(float(v))
                except (TypeError, ValueError):
                    raise ValueError(f"数值字段无效: {num_field}={mapped[num_field]}")
            else:
                mapped[num_field] = 0 if num_field != "conversion_value" else 0.0
        mapped["spend"] = round(float(mapped["spend"]), 2)
        mapped["conversion_value"] = round(float(mapped["conversion_value"]), 2)

        # 全 0 记录视为无效行
        if (mapped["impressions"] <= 0 and mapped["clicks"] <= 0
                and mapped["conversions"] <= 0 and mapped["spend"] <= 0
                and mapped["conversion_value"] <= 0):
            raise ValueError("该行无有效数值（全为 0）")

        # 双轨记录
        rec = {"raw": raw_clean, "mapped": mapped, "tags": {}}
        if raw.get("sample"):
            rec["sample"] = True
        return rec

    @staticmethod
    def _parse_date(val: str) -> date:
        """解析日期，支持 YYYY-MM-DD / YYYY/MM/DD / 时间戳（秒/毫秒）"""
        val = val.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
        # 时间戳
        try:
            ts = float(val)
            if ts > 1e12:
                ts /= 1000
            return datetime.fromtimestamp(ts).date()
        except ValueError:
            pass
        raise ValueError(f"无法解析日期: {val}")

    # ==================== CSV / JSON 解析 ====================
    def parse_csv(self, content: bytes) -> list:
        """解析 CSV 内容为记录列表"""
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise ValueError("CSV 缺少表头")
        return [dict(row) for row in reader]

    def parse_json(self, content: bytes) -> list:
        """解析 JSON 内容（数组）为记录列表"""
        text = content.decode("utf-8-sig", errors="replace")
        data = json.loads(text)
        if isinstance(data, dict):
            # 兼容 {"records": [...]}
            data = data.get("records", [])
        if not isinstance(data, list):
            raise ValueError("JSON 应为记录数组")
        return data

    # ==================== 指标计算 ====================
    @staticmethod
    def _calc(records: list) -> dict:
        """聚合计算关键指标（对给定记录集合，读 mapped 域）"""
        def get(r, k, default=0):
            v = r.get("mapped", {}).get(k, default)
            try:
                return float(v) if k in ("spend", "conversion_value") else int(v)
            except (TypeError, ValueError):
                return default
        impressions = sum(get(r, "impressions") for r in records)
        clicks = sum(get(r, "clicks") for r in records)
        conversions = sum(get(r, "conversions") for r in records)
        spend = sum(get(r, "spend", 0.0) for r in records)
        conv_value = sum(get(r, "conversion_value", 0.0) for r in records)
        return {
            "impressions": impressions,
            "clicks": clicks,
            "conversions": conversions,
            "spend": round(spend, 2),
            "conversion_value": round(conv_value, 2),
            "ctr": round(clicks / impressions * 100, 2) if impressions else 0.0,
            "cvr": round(conversions / clicks * 100, 2) if clicks else 0.0,
            "cpc": round(spend / clicks, 2) if clicks else 0.0,
            "cpm": round(spend / impressions * 1000, 2) if impressions else 0.0,
            "cpa": round(spend / conversions, 2) if conversions else 0.0,
            "roas": round(conv_value / spend, 2) if spend else 0.0,
        }

    def summary(self, account: Optional[str] = None, days: Optional[int] = None) -> dict:
        """投放数据总览：关键指标 + 健康评分 + 信号统计"""
        records = self._filter(account=account, days=days)
        metrics = self._calc(records)
        signals = self.detect_signals(records)
        health = self._health_score(signals)
        accounts = len({r.get("mapped", {}).get("account") for r in records})
        metrics.update({
            "account_count": accounts,
            "day_count": len({r.get("mapped", {}).get("date") for r in records}),
            "record_count": len(records),
        })
        return {
            "metrics": metrics,
            "health": health,
            "anomaly_count": len(signals),
            "anomaly_by_severity": {
                sev: len([a for a in signals if a["severity"] == sev])
                for sev in (SEV_CRITICAL, SEV_HIGH, SEV_MEDIUM, SEV_LOW)
            },
        }

    def trend(self, account: Optional[str] = None, days: Optional[int] = None) -> list:
        """时间维度趋势：按日聚合，返回每日指标序列"""
        records = self._filter(account=account, days=days)
        by_date = defaultdict(list)
        for r in records:
            d = r.get("mapped", {}).get("date")
            if d:
                by_date[d].append(r)
        trend = []
        for d in sorted(by_date.keys()):
            daily = self._calc(by_date[d])
            daily["date"] = d
            trend.append(daily)
        return trend

    def by_account(self, days: Optional[int] = None) -> list:
        """账户维度对比：按账户聚合全部指标"""
        records = self._filter(days=days)
        by_acc = defaultdict(list)
        for r in records:
            acc = r.get("mapped", {}).get("account")
            if acc:
                by_acc[acc].append(r)
        result = []
        for acc, recs in sorted(by_acc.items()):
            m = self._calc(recs)
            m["account"] = acc
            m["day_count"] = len({r.get("mapped", {}).get("date") for r in recs})
            m["record_count"] = len(recs)
            result.append(m)
        return result

    # ==================== 信号规则引擎（Phase3） ====================

    @property
    def _rule_state_file(self) -> Path:
        return settings.audit_rule_state_file

    def _load_rule_state(self) -> dict:
        """加载用户规则启用状态（持久化），与 spec 默认合并
        状态格式：{rule_id: {"enabled": bool, "method": "window"|"daily"}}
        兼容旧格式：{rule_id: bool}
        """
        saved = {}
        if self._rule_state_file.exists():
            try:
                saved = json.loads(self._rule_state_file.read_text(encoding="utf-8"))
            except (json.JSONDecodeError, OSError):
                saved = {}
        rules = settings.audit_signal_rules
        state = {}
        for rid, rcfg in rules.items():
            default_enabled = bool(rcfg.get("default_enabled", True))
            default_method = rcfg.get("method", "window")
            entry = saved.get(rid)
            if isinstance(entry, dict):
                state[rid] = {
                    "enabled": bool(entry.get("enabled", default_enabled)),
                    "method": entry.get("method", default_method),
                }
            else:
                # 旧格式：直接 bool
                state[rid] = {
                    "enabled": default_enabled if entry is None else bool(entry),
                    "method": default_method,
                }
        return state

    def _save_rule_state(self, state: dict):
        self._rule_state_file.parent.mkdir(parents=True, exist_ok=True)
        self._rule_state_file.write_text(
            json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8"
        )

    def get_rule_states(self) -> dict:
        """获取全部信号规则定义 + 启用状态 + 检测方法"""
        rules = settings.audit_signal_rules
        state = self._load_rule_state()
        result = {}
        for rid, rcfg in rules.items():
            entry = state.get(rid, {})
            result[rid] = {
                "id": rid,
                "name": rcfg.get("name", rid),
                "description": rcfg.get("description", ""),
                "category": rcfg.get("category", "hint"),
                "enabled": entry.get("enabled", True),
                "method": entry.get("method", rcfg.get("method", "window")),
                "methods": ["window", "daily"] if rcfg.get("method") else [],
                "default_enabled": bool(rcfg.get("default_enabled", True)),
            }
        return result

    def set_rule_state(self, rule_id: str, enabled: bool, method: Optional[str] = None) -> dict:
        """设置单个规则启用状态（可选更新检测方法）"""
        rules = settings.audit_signal_rules
        if rule_id not in rules:
            raise ValueError(f"未知规则: {rule_id}")
        if method is not None and method not in ("window", "daily"):
            raise ValueError(f"未知检测方法: {method}（应为 window 或 daily）")
        state = self._load_rule_state()
        entry = state.setdefault(rule_id, {})
        entry["enabled"] = bool(enabled)
        if method:
            entry["method"] = method
        self._save_rule_state(state)
        return {"ok": True, "rule_id": rule_id, "enabled": bool(enabled), "method": entry.get("method")}

    def reset_rule_states(self) -> dict:
        """重置全部规则为默认启用"""
        self._save_rule_state({})
        return {"ok": True}

    def _is_rule_enabled(self, rule_id: str) -> bool:
        state = self._load_rule_state()
        return state.get(rule_id, {}).get("enabled", True) if isinstance(state.get(rule_id), dict) else bool(state.get(rule_id, True))

    # ---- 统计工具（MAD 抗离群） ----
    @staticmethod
    def _mad_zscore(values: list, x: float) -> float:
        """计算 x 相对 values 的稳健 z-score（中位数 + MAD）
        z = (x - median) / (1.4826 * MAD)
        """
        if not values:
            return 0.0
        median = statistics.median(values)
        if len(values) >= 2:
            deviations = [abs(v - median) for v in values]
            mad = statistics.median(deviations)
        else:
            mad = 0.0
        if mad == 0:
            # MAD 为 0（值都相等）时，用均值 + 小 epsilon 兜底
            mean = sum(values) / len(values)
            if mean == 0:
                return 0.0
            scale = max(mean * 0.05, 1e-9)
        else:
            scale = 1.4826 * mad
        return (x - median) / scale

    def _mk_signal(self, severity, category, title, description, suggestion="",
                   account=None, date=None, impact_amount=0.0, impact_desc="",
                   metrics=None) -> dict:
        """构造统一 schema 的信号"""
        return {
            "severity": severity,
            "category": category,
            "title": title,
            "description": description,
            "suggestion": suggestion,
            "account": account,
            "date": date,
            "impact": {"amount": round(impact_amount, 2), "desc": impact_desc},
            "metrics": metrics or {},
        }

    def detect_signals(self, records: Optional[list] = None) -> list:
        """信号引擎主入口：合并 6 类规则信号（参照 Claude-ads + marqops 业界实践）
        统一 schema：severity / category / impact / metrics
        规则启用状态可配置（默认全部启用，健康评分默认启用）
        """
        records = self._filter() if records is None else records
        if not records:
            return [self._mk_signal(
                SEV_LOW, "hint", "暂无投放数据",
                "上传 CSV/JSON 数据或生成示例数据后开始审计",
                suggestion="前往「数据导入」上传投放数据",
                impact_desc="无数据", metrics={"record_count": 0},
            )]
        signals = []
        rules = settings.audit_signal_rules
        state = self._load_rule_state()
        # 展开为 {rule_id: enabled_bool}，便于判断
        enabled_map = {rid: entry.get("enabled", True) for rid, entry in state.items()}

        # 全局聚合
        total = self._calc(records)
        dates_all = sorted({r.get("mapped", {}).get("date", "") for r in records})

        # ---- 提示类：样本不足 / 曝光不足 ----
        if enabled_map.get("small_sample", True):
            cfg = rules.get("small_sample", {})
            min_days = int(cfg.get("min_sample_days", 14))
            if len(dates_all) < min_days:
                signals.append(self._mk_signal(
                    SEV_LOW, "hint", "数据样本不足",
                    f"仅有 {len(dates_all)} 天数据，趋势分析参考价值有限",
                    "建议补充至少 14 天的投放数据",
                    impact_desc="统计可靠性不足",
                    metrics={"days": len(dates_all), "min_days": min_days},
                ))
        if enabled_map.get("low_impressions", True):
            cfg = rules.get("low_impressions", {})
            min_imp = int(cfg.get("min_impressions", 10000))
            if total["impressions"] < min_imp:
                signals.append(self._mk_signal(
                    SEV_LOW, "hint", "曝光量不足",
                    f"总曝光 {total['impressions']:,}，样本可能不足以支撑可靠结论",
                    "建议补充更长周期或更多账户的数据",
                    impact_desc="样本量偏小",
                    metrics={"impressions": total["impressions"], "min": min_imp},
                ))

        # ---- 按账户时间序列（花费突增 / CTR 骤降 / 连续无转化） ----
        for acc, acc_recs in self._group_by_account(records).items():
            dates = sorted({r.get("mapped", {}).get("date", "") for r in acc_recs})
            day_map = {d: self._calc([r for r in acc_recs if r.get("mapped", {}).get("date") == d]) for d in dates}

            # 花费突增（按 method 分发：window 窗口累计法 / daily 逐日 MAD）
            if enabled_map.get("spend_surge", True):
                cfg = dict(rules.get("spend_surge", {}))
                method = state.get("spend_surge", {}).get("method") or cfg.get("method", "window")
                if method == "daily":
                    signals += self._detect_spend_surge_daily(acc, dates, day_map, cfg)
                else:
                    signals += self._detect_spend_surge_window(acc, dates, day_map, cfg)

            # CTR 骤降（按 method 分发）
            if enabled_map.get("ctr_drop", True):
                cfg = dict(rules.get("ctr_drop", {}))
                method = state.get("ctr_drop", {}).get("method") or cfg.get("method", "window")
                if method == "daily":
                    signals += self._detect_ctr_drop_daily(acc, dates, day_map, cfg)
                else:
                    signals += self._detect_ctr_drop_window(acc, dates, day_map, cfg)

            # 连续花费无转化（数据质量分流：追踪中断 vs 业务下滑）
            if enabled_map.get("no_conversion", True):
                signals += self._detect_no_conversion(acc, dates, day_map, rules.get("no_conversion", {}))

        # ---- 账户对比（ROAS 过低 / CPA 过高） ----
        if enabled_map.get("roas_low", True):
            signals += self._detect_roas_low(records, total, rules.get("roas_low", {}))
        if enabled_map.get("cpa_high", True):
            signals += self._detect_cpa_high(records, total, rules.get("cpa_high", {}))

        # ---- 标签组合权重变化（Phase4，需打标数据） ----
        if enabled_map.get("weight_shift", True):
            signals += self._detect_weight_shift(records, rules.get("weight_shift", {}))

        # ---- 排序：severity + impact 金额降序 ----
        sev_order = {SEV_CRITICAL: 0, SEV_HIGH: 1, SEV_MEDIUM: 2, SEV_LOW: 3}
        signals.sort(key=lambda s: (sev_order.get(s["severity"], 9), -s["impact"]["amount"]))
        return signals

    # ---- 各规则检测实现 ----

    def _detect_spend_surge_window(self, acc, dates, day_map, cfg) -> list:
        """花费突增：窗口累计法（业界实践，默认）
        最近 N 天（duration_days）累计花费 vs 历史日均中位数 × N：
          - 累计超额 = 窗口累计 - 历史日均 × 窗口
          - 触发：超额比例 ≥ min_change_pct 且 超额金额 ≥ min_impact
        非重叠窗口去重：同一账户检测到一次后跳过 window 天，避免滑动窗口重复报警。
        """
        signals = []
        min_pct = float(cfg.get("min_change_pct", 50.0))
        min_impact = float(cfg.get("min_impact_amount", 2000.0))
        window = max(int(cfg.get("duration_days", 2)), 1)
        last_trigger_idx = -window  # 上次触发位置（非重叠窗口）
        for i in range(len(dates)):
            if i < window:
                continue
            # 跳过上次触发附近的窗口（非重叠去重）
            if i - last_trigger_idx < window:
                continue
            # 最近 window 天窗口（含当日）
            cur_total = sum(day_map[x]["spend"] for x in dates[i - window + 1:i + 1])
            # 窗口之前的历史日均（中位数，抗离群）
            hist = [day_map[x]["spend"] for x in dates[:i - window + 1]]
            if not hist:
                continue
            median_daily = statistics.median(hist)
            if median_daily <= 0:
                continue
            expected = median_daily * window
            excess = cur_total - expected
            pct_change = excess / expected * 100
            if pct_change < min_pct or excess < min_impact:
                continue
            ratio = cur_total / expected
            severity = SEV_HIGH if ratio >= 2.0 else SEV_MEDIUM
            signals.append(self._mk_signal(
                severity, "surge", f"花费突增（{ratio:.1f} 倍）",
                f"「{acc}」最近 {window} 天（{dates[i - window + 1]}~{dates[i]}）花费 ¥{cur_total:,.2f}，"
                f"高于历史日均 ¥{median_daily:,.2f} × {window} 天预期的 {pct_change:.0f}%",
                "检查是否误配预算/出价、流量质量下降或异常点击",
                account=acc, date=dates[i], impact_amount=excess,
                impact_desc="预估超额花费",
                metrics={"ratio": round(ratio, 2), "magnitude": f"{pct_change:.0f}%",
                         "window_days": window, "excess": round(excess, 2)},
            ))
            last_trigger_idx = i
        return signals

    def _detect_ctr_drop_window(self, acc, dates, day_map, cfg) -> list:
        """CTR 骤降：窗口累计法（业界实践，默认）
        最近 N 天窗口 CTR（累计点击/累计曝光）vs 历史窗口 CTR 中位数：
          触发条件：下降比例 ≥ min_change_pct 且 曝光足够 且 预估损失 ≥ min_impact
        非重叠窗口去重。
        """
        signals = []
        min_pct = float(cfg.get("min_change_pct", 40.0))
        min_imp = float(cfg.get("min_impressions", 10000))
        min_impact = float(cfg.get("min_impact_amount", 1000.0))
        window = max(int(cfg.get("duration_days", 2)), 1)
        last_trigger_idx = -window
        for i in range(len(dates)):
            if i < window:
                continue
            if i - last_trigger_idx < window:
                continue
            # 窗口累计点击/曝光 → 窗口 CTR
            win_dates = dates[i - window + 1:i + 1]
            win_clicks = sum(day_map[x]["clicks"] for x in win_dates)
            win_imp = sum(day_map[x]["impressions"] for x in win_dates)
            win_spend = sum(day_map[x]["spend"] for x in win_dates)
            if win_imp < min_imp:
                continue
            win_ctr = win_clicks / win_imp * 100 if win_imp > 0 else 0
            if win_ctr <= 0:
                continue
            # 历史窗口 CTR（每个前 window 天窗口计算一次，取中位数）
            hist_windows = []
            for j in range(window, i - window + 1):
                h_dates = dates[j - window + 1:j + 1]
                h_clicks = sum(day_map[x]["clicks"] for x in h_dates)
                h_imp = sum(day_map[x]["impressions"] for x in h_dates)
                if h_imp >= min_imp:
                    hist_windows.append(h_clicks / h_imp * 100)
            if not hist_windows:
                continue
            median_ctr = statistics.median(hist_windows)
            if median_ctr <= 0:
                continue
            pct_drop = (median_ctr - win_ctr) / median_ctr * 100
            if pct_drop < min_pct:
                continue
            # 预估损失 = 曝光不变时损失的点击 × 平均 CPC（曝光加权，避免花费随点击下降而低估）
            # 损失点击 = win_imp × (median_ctr - win_ctr) / 100
            lost_clicks = win_imp * (median_ctr - win_ctr) / 100
            avg_cpc = win_spend / win_clicks if win_clicks > 0 else 0
            impact = lost_clicks * avg_cpc
            if impact < min_impact:
                continue
            severity = SEV_HIGH if pct_drop >= min_pct * 2 else SEV_MEDIUM
            signals.append(self._mk_signal(
                severity, "decay", f"点击率骤降（{pct_drop:.0f}%）",
                f"「{acc}」最近 {window} 天（{win_dates[0]}~{win_dates[-1]}）窗口 CTR {win_ctr:.2f}%，"
                f"低于历史窗口 CTR 中位 {median_ctr:.2f}% 达 {pct_drop:.0f}%",
                "检查素材疲劳、受众定向变化或展示位置质量",
                account=acc, date=win_dates[-1], impact_amount=impact,
                impact_desc="预估因 CTR 下降损失",
                metrics={"magnitude": f"{pct_drop:.0f}%", "win_ctr": round(win_ctr, 2),
                         "median_ctr": round(median_ctr, 2), "window_days": window},
            ))
            last_trigger_idx = i
        return signals

    def _detect_no_conversion(self, acc, dates, day_map, cfg) -> list:
        """连续花费无转化：数据质量分流（追踪中断 vs 业务下滑）"""
        signals = []
        no_conv_days = int(cfg.get("no_conversion_days", 3))
        min_impact = float(cfg.get("min_impact_amount", 500.0))
        imp_ratio = float(cfg.get("tracking_break_imp_ratio", 0.7))
        streak = 0
        for d in dates:
            dm = day_map[d]
            if dm["spend"] > 0 and dm["conversions"] == 0:
                streak += 1
                if streak == no_conv_days:
                    impact = dm["spend"]
                    # 分流：检查曝光是否相对历史正常（正常 → 追踪中断；下降 → 业务下滑）
                    imp_now = dm["impressions"]
                    idx = dates.index(d)
                    hist_imp = [day_map[x]["impressions"] for x in dates[:idx]] if idx > 0 else []
                    if hist_imp:
                        median_imp = statistics.median(hist_imp)
                        is_tracking_break = (median_imp > 0 and imp_now >= median_imp * imp_ratio)
                    else:
                        is_tracking_break = True  # 无历史则归为数据质量
                    if impact < min_impact and not is_tracking_break:
                        streak = 0
                        continue
                    if is_tracking_break:
                        signals.append(self._mk_signal(
                            SEV_HIGH, "data_quality", "连续花费无转化（疑似追踪中断）",
                            f"「{acc}」连续 {streak} 天有花费但转化量为 0，曝光 {imp_now:,} 相对历史正常，可能为转化追踪中断而非业务下滑",
                            "检查 Pixel/转化 API/落地页埋点是否正常",
                            account=acc, date=d, impact_amount=impact,
                            impact_desc="连续无转化期间的累计花费",
                            metrics={"duration_days": streak, "impressions": imp_now,
                                     "tracking_break": True},
                        ))
                    else:
                        signals.append(self._mk_signal(
                            SEV_HIGH, "decay", "连续花费无转化（业务下滑）",
                            f"「{acc}」连续 {streak} 天有花费但转化量为 0，曝光同步下降",
                            "暂停该账户/系列并排查受众、素材与落地页",
                            account=acc, date=d, impact_amount=impact,
                            impact_desc="连续无转化期间的累计花费",
                            metrics={"duration_days": streak, "impressions": imp_now,
                                     "tracking_break": False},
                        ))
            else:
                streak = 0
        return signals

    # ---- 逐日法（daily）：单日 MAD z-score，对单日突变敏感 ----
    def _detect_spend_surge_daily(self, acc, dates, day_map, cfg) -> list:
        """花费突增：逐日法
        每日 MAD 稳健 z-score（同周几样本≥3 时用同周几，否则连续历史）：
          - z ≥ z_threshold 且 变化幅度 ≥ min_change_pct 且 超额 ≥ min_impact
          - 连续 duration_days 天达标才触发（防单日噪声）
        """
        signals = []
        z_th = float(cfg.get("z_threshold", 3.0))
        z_hi = float(cfg.get("z_high_threshold", 5.0))
        min_pct = float(cfg.get("min_change_pct", 100.0))
        min_impact = float(cfg.get("min_impact_amount", 5000.0))
        duration = max(int(cfg.get("duration_days", 2)), 1)
        wd_win = int(cfg.get("weekday_window", 4))
        streak = 0
        last_trigger_idx = -duration
        for i, d in enumerate(dates):
            if i < 1:
                continue
            if i - last_trigger_idx < duration:
                continue
            cur = day_map[d]
            spend = cur["spend"]
            same_wd = [day_map[x]["spend"] for x in dates[:i]
                       if datetime.strptime(x, "%Y-%m-%d").weekday() == datetime.strptime(d, "%Y-%m-%d").weekday()]
            hist = same_wd if wd_win > 0 and len(same_wd) >= 3 else [day_map[x]["spend"] for x in dates[:i]]
            if not hist:
                continue
            z = self._mad_zscore(hist, spend)
            median = statistics.median(hist)
            pct_change = (spend - median) / median * 100 if median > 0 else 0
            excess = spend - median
            is_surge = (z >= z_th and pct_change >= min_pct and excess >= min_impact)
            if is_surge:
                streak += 1
                if streak >= duration:
                    severity = SEV_HIGH if z >= z_hi else SEV_MEDIUM
                    signals.append(self._mk_signal(
                        severity, "surge", f"花费突增（z={z:.1f}）",
                        f"「{acc}」{d} 花费 ¥{spend:,.2f}，相对历史中位 ¥{median:,.2f} 上涨 {pct_change:.0f}%（逐日 MAD z={z:.1f}，持续 {streak} 天）",
                        "检查是否误配预算/出价、流量质量下降或异常点击",
                        account=acc, date=d, impact_amount=excess,
                        impact_desc="预估超额花费",
                        metrics={"z_score": round(z, 2), "magnitude": f"{pct_change:.0f}%",
                                 "median": round(median, 2), "duration_days": streak, "method": "daily"},
                    ))
                    last_trigger_idx = i
            else:
                streak = 0
        return signals

    def _detect_ctr_drop_daily(self, acc, dates, day_map, cfg) -> list:
        """CTR 骤降：逐日法（每日 MAD z-score，对单日变化敏感）"""
        signals = []
        z_th = float(cfg.get("z_threshold", 3.0))
        z_hi = float(cfg.get("z_high_threshold", 5.0))
        min_pct = float(cfg.get("min_change_pct", 50.0))
        min_imp = float(cfg.get("min_impressions", 10000))
        min_impact = float(cfg.get("min_impact_amount", 2000.0))
        duration = max(int(cfg.get("duration_days", 2)), 1)
        wd_win = int(cfg.get("weekday_window", 4))
        streak = 0
        last_trigger_idx = -duration
        for i, d in enumerate(dates):
            if i < 1:
                continue
            if i - last_trigger_idx < duration:
                continue
            cur = day_map[d]
            if cur["impressions"] < min_imp or cur["ctr"] <= 0:
                streak = 0
                continue
            same_wd = [day_map[x]["ctr"] for x in dates[:i]
                       if datetime.strptime(x, "%Y-%m-%d").weekday() == datetime.strptime(d, "%Y-%m-%d").weekday()]
            hist = same_wd if wd_win > 0 and len(same_wd) >= 3 else [day_map[x]["ctr"] for x in dates[:i]]
            if not hist:
                continue
            z = self._mad_zscore(hist, cur["ctr"])
            median = statistics.median(hist)
            pct_drop = (median - cur["ctr"]) / median * 100 if median > 0 else 0
            lost_clicks = cur["impressions"] * (median - cur["ctr"]) / 100
            avg_cpc = cur["spend"] / cur["clicks"] if cur["clicks"] > 0 else 0
            impact = lost_clicks * avg_cpc
            is_drop = (z <= -z_th and pct_drop >= min_pct and impact >= min_impact)
            if is_drop:
                streak += 1
                if streak >= duration:
                    severity = SEV_HIGH if z <= -z_hi else SEV_MEDIUM
                    signals.append(self._mk_signal(
                        severity, "decay", f"点击率骤降（{pct_drop:.0f}%）",
                        f"「{acc}」{d} CTR {cur['ctr']}%，相对历史中位 {median}% 下降 {pct_drop:.0f}%（逐日 MAD z={z:.1f}，持续 {streak} 天）",
                        "检查素材疲劳、受众定向变化或展示位置质量",
                        account=acc, date=d, impact_amount=impact,
                        impact_desc="预估因 CTR 下降损失",
                        metrics={"z_score": round(z, 2), "magnitude": f"{pct_drop:.0f}%",
                                 "median": median, "duration_days": streak, "method": "daily"},
                    ))
                    last_trigger_idx = i
            else:
                streak = 0
        return signals

    # ---- 标签组合权重变化（Phase4） ----
    def _detect_weight_shift(self, records, cfg) -> list:
        """权重变化信号：某标签组合在本批数据中花费占比显著变化
        近期窗口（recent_days）内组合花费占比 vs 历史基线窗口（baseline_days）占比：
          - 占比变化绝对值 ≥ min_share_change 且 组合占比 ≥ min_share 才触发
          - 提示「策略调整迹象」
        """
        signals = []
        recent_days = int(cfg.get("recent_days", 7))
        baseline_days = int(cfg.get("baseline_days", 14))
        min_share_change = float(cfg.get("min_share_change", 0.2))
        min_share = float(cfg.get("min_share", 0.05))
        min_impact = float(cfg.get("min_impact_amount", 1000.0))

        # 只有打标数据才参与
        tagged = [r for r in records if r.get("tags")]
        if not tagged or len(tagged) < 10:
            return signals

        today = date.today()
        recent_cutoff = (today - timedelta(days=recent_days)).isoformat()
        baseline_cutoff = (today - timedelta(days=baseline_days)).isoformat()

        def share_by_combo(recs, period_start, period_end=None):
            """计算每个标签组合的花费占比（组合 = 全部标签组值拼接）"""
            combo_spend = defaultdict(float)
            total_spend = 0.0
            for r in recs:
                d = r.get("mapped", {}).get("date", "")
                if d < period_start:
                    continue
                if period_end and d > period_end:
                    continue
                tags = r.get("tags", {})
                # 组合键：排序后的所有标签值拼接（含账户）
                combo_parts = [r.get("mapped", {}).get("account", "未分组")]
                for gid in sorted(tags.keys()):
                    combo_parts.extend(sorted(tags[gid]))
                key = " · ".join(combo_parts)
                spend = float(r.get("mapped", {}).get("spend", 0))
                combo_spend[key] += spend
                total_spend += spend
            shares = {}
            if total_spend > 0:
                for k, v in combo_spend.items():
                    shares[k] = v / total_spend
            return shares, total_spend

        # 近期占比 vs 基线占比
        recent_share, recent_total = share_by_combo(tagged, recent_cutoff)
        baseline_share, baseline_total = share_by_combo(tagged, baseline_cutoff, recent_cutoff)
        if not recent_share or not baseline_share:
            return signals

        # 检测变化（近期上升的组合）
        for combo, recent_pct in recent_share.items():
            if recent_pct < min_share:
                continue
            base_pct = baseline_share.get(combo, 0.0)
            change = recent_pct - base_pct
            if abs(change) < min_share_change:
                continue
            impact = recent_total * abs(change)  # 占比变化 × 近期总花费
            if impact < min_impact:
                continue
            direction = "上升" if change > 0 else "下降"
            severity = SEV_MEDIUM if abs(change) >= min_share_change * 1.5 else SEV_LOW
            signals.append(self._mk_signal(
                severity, "shift", f"组合权重{direction}（{abs(change) * 100:.0f}%）",
                f"组合「{combo}」花费占比从基线期 {base_pct * 100:.0f}% 变化至近期 {recent_pct * 100:.0f}%"
                f"（{direction} {abs(change) * 100:.0f}%）",
                "关注该组合的投放策略调整迹象，验证是否有意为之",
                impact_amount=impact, impact_desc="占比变化对应的预算规模",
                metrics={"recent_share": round(recent_pct, 3), "baseline_share": round(base_pct, 3),
                         "change": round(change, 3), "recent_days": recent_days},
            ))
        return signals

    # ==================== 多维透视（Phase4） ====================
    def pivot(self, dimensions: list, days: Optional[int] = None,
              account: Optional[str] = None, metric: str = "spend") -> dict:
        """多维透视：按标签组组合聚合
        dimensions: 标签组 ID 列表（如 ["creative_keyword", "audience"]）
        返回每个组合的 sum/avg 指标 + 记录数
        """
        if not dimensions:
            return {"combinations": [], "total": {}}
        records = self._filter(account=account, days=days)
        tagged = [r for r in records if r.get("tags")]
        if not tagged:
            return {"combinations": [], "total": self._calc(records)}

        # 维度名映射
        lib = self._load_tag_lib()
        group_names = {g["id"]: g["name"] for g in lib.get("groups", [])}

        combo_map = defaultdict(list)
        for r in tagged:
            tags = r.get("tags", {})
            # 组合键：每个维度的第一个标签值（或「未标注」）
            key_parts = []
            for gid in dimensions:
                vals = tags.get(gid, [])
                key_parts.append(vals[0] if vals else "未标注")
            combo_map[tuple(key_parts)].append(r)

        combinations = []
        for key, recs in combo_map.items():
            m = self._calc(recs)
            m["dimensions"] = dict(zip(dimensions, key))
            m["dimension_labels"] = {gid: group_names.get(gid, gid) for gid in dimensions}
            m["record_count"] = len(recs)
            combinations.append(m)

        # 按 metric 排序（默认 spend 降序）
        combinations.sort(key=lambda c: c.get(metric, 0), reverse=True)
        return {
            "combinations": combinations,
            "total": self._calc(tagged),
            "dimensions": dimensions,
            "dimension_labels": {gid: group_names.get(gid, gid) for gid in dimensions},
            "metric": metric,
        }

    # ==================== Excel 导出（Phase5） ====================
    def export_excel(self, account: Optional[str] = None, days: Optional[int] = None,
                     dimensions: Optional[list] = None) -> bytes:
        """导出结构化 Excel（xlsx）：
        Sheet1「原始数据」：raw 全字段 + 打标列展开（每标签组一列）+ mapped 派生指标
        Sheet2「多维透视」：按标签组合聚合（若指定 dimensions 或有打标数据）
        Sheet3「数据信号」：当前信号列表（含 impact/metrics）
        返回 xlsx 文件字节
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        records = self._filter(account=account, days=days)
        wb = Workbook()

        # ---- Sheet1: 原始数据 ----
        ws = wb.active
        ws.title = "原始数据"
        # 收集列：raw 字段 + 标签组列 + 派生指标
        lib = self._load_tag_lib()
        group_names = {g["id"]: g["name"] for g in lib.get("groups", [])}
        raw_fields = []
        seen = set()
        for r in records:
            for k in r.get("raw", {}).keys():
                if k not in seen:
                    seen.add(k)
                    raw_fields.append(k)
        tag_group_ids = list(group_names.keys())
        derived = ["CTR%", "CVR%", "CPC", "CPM", "CPA", "ROAS"]

        header = raw_fields + [group_names.get(gid, gid) for gid in tag_group_ids] + derived
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F46E5")
        ws.append(header)
        for c in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r in records:
            row = []
            for f in raw_fields:
                row.append(r.get("raw", {}).get(f, ""))
            for gid in tag_group_ids:
                tags = r.get("tags", {}).get(gid, [])
                row.append("、".join(tags) if tags else "")
            m = r.get("mapped", {})
            row.extend([
                m.get("ctr", 0), m.get("cvr", 0), m.get("cpc", 0),
                m.get("cpm", 0), m.get("cpa", 0), m.get("roas", 0),
            ])
            ws.append(row)
        # 列宽自适应（粗略）
        for c in range(1, len(header) + 1):
            ws.column_dimensions[get_column_letter(c)].width = max(12, min(30, len(str(header[c - 1])) * 2 + 4))

        # ---- Sheet2: 多维透视 ----
        ws2 = wb.create_sheet("多维透视")
        tagged = [r for r in records if r.get("tags")]
        ws2.append(["维度组合", "记录数", "花费", "转化", "曝光", "点击", "转化价值", "CTR%", "CPA", "ROAS"])
        for c in range(1, 11):
            ws2.cell(row=1, column=c).font = header_font
            ws2.cell(row=1, column=c).fill = header_fill
        if tagged:
            dims = dimensions or list(group_names.keys())[:1]
            combo_map = defaultdict(list)
            for r in tagged:
                tags = r.get("tags", {})
                key_parts = []
                for gid in dims:
                    vals = tags.get(gid, [])
                    key_parts.append(vals[0] if vals else "未标注")
                combo_map[tuple(key_parts)].append(r)
            for key, recs in combo_map.items():
                m = self._calc(recs)
                ws2.append([
                    " + ".join(key), len(recs),
                    round(m["spend"], 2), m["conversions"], m["impressions"], m["clicks"],
                    round(m["conversion_value"], 2), m["ctr"], m["cpa"], m["roas"],
                ])
        else:
            ws2.append(["暂无打标数据，无法透视"])

        # ---- Sheet3: 数据信号 ----
        ws3 = wb.create_sheet("数据信号")
        signals = self.detect_signals(records)
        ws3.append(["严重度", "分类", "标题", "描述", "影响金额", "影响说明", "账户", "日期", "指标详情"])
        for c in range(1, 10):
            ws3.cell(row=1, column=c).font = header_font
            ws3.cell(row=1, column=c).fill = header_fill
        for s in signals:
            ws3.append([
                s["severity"], s["category"], s["title"], s["description"],
                s["impact"]["amount"], s["impact"]["desc"],
                s.get("account", ""), s.get("date", ""),
                json.dumps(s.get("metrics", {}), ensure_ascii=False),
            ])

        # 导出到内存
        import io
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def _detect_roas_low(self, records, total, cfg) -> list:
        """ROAS 过低：账户维度，持续低于警戒线"""
        signals = []
        roas_warn = float(cfg.get("roas_warn_below", 1.0))
        min_impact = float(cfg.get("min_impact_amount", 1000.0))
        for acc, acc_recs in self._group_by_account(records).items():
            m = self._calc(acc_recs)
            if m["spend"] > 0 and m["roas"] < roas_warn:
                impact = m["spend"] * (roas_warn - m["roas"])  # 与警戒线差距的预估浪费
                if impact < min_impact:
                    continue
                signals.append(self._mk_signal(
                    SEV_HIGH if m["spend"] > 5000 else SEV_MEDIUM,
                    "inefficiency", "投产比低于警戒线",
                    f"「{acc}」ROAS {m['roas']} < {roas_warn}，花费 ¥{m['spend']:,.2f} 未产生足够转化价值",
                    "评估是否暂停该账户或大幅调整预算分配",
                    account=acc, impact_amount=impact,
                    impact_desc="相对警戒线的预估投入损失",
                    metrics={"roas": m["roas"], "spend": m["spend"], "warn_below": roas_warn},
                ))
        return signals

    def _detect_cpa_high(self, records, total, cfg) -> list:
        """CPA 过高：账户 CPA 显著高于整体均值"""
        signals = []
        ratio = float(cfg.get("cpa_surge_ratio", 2.0))
        min_impact = float(cfg.get("min_impact_amount", 0.0))
        total_cpa = total["cpa"]
        if total_cpa <= 0:
            return signals
        for acc, acc_recs in self._group_by_account(records).items():
            m = self._calc(acc_recs)
            if m["cpa"] > total_cpa * ratio and m["conversions"] > 0:
                impact = (m["cpa"] - total_cpa) * m["conversions"]  # 超额成本 × 转化数
                if impact < min_impact:
                    continue
                signals.append(self._mk_signal(
                    SEV_MEDIUM, "inefficiency", "获客成本过高",
                    f"「{acc}」CPA ¥{m['cpa']:,.2f}，为整体均值（¥{total_cpa:,.2f}）的 {m['cpa'] / total_cpa:.1f} 倍",
                    "拆分优化该账户的关键词/素材/出价策略",
                    account=acc, impact_amount=impact,
                    impact_desc="相对整体均值的超额获客成本",
                    metrics={"cpa": m["cpa"], "avg_cpa": total_cpa, "ratio": round(m["cpa"] / total_cpa, 1)},
                ))
        return signals

    # ---- 兼容旧 API ----
    def detect_anomalies(self, records: Optional[list] = None) -> list:
        """兼容旧接口：调用新信号引擎，映射旧字段名（type/suggestion）"""
        signals = self.detect_signals(records)
        result = []
        for s in signals:
            result.append({
                "severity": s["severity"],
                "type": s["category"],
                "title": s["title"],
                "description": s["description"],
                "suggestion": s["suggestion"],
                "date": s["date"],
                "account": s["account"],
            })
        return result

    def _health_score(self, anomalies: list) -> dict:
        """健康评分（参照 Claude-ads：基础 100，按严重度扣分，A/B/C/D 分级）"""
        cfg = settings.audit_health
        score = 100
        for a in anomalies:
            score += int(cfg.get(a["severity"], 0))
        score = max(0, min(100, score))
        grades = cfg.get("grades", ["A", "B", "C", "D"])
        thresholds = cfg.get("grade_thresholds", [90, 75, 60, 0])
        grade = grades[0]
        for g, t in zip(grades, thresholds):
            if score >= t:
                grade = g
                break
        return {"score": score, "grade": grade}

    # ==================== 工具 ====================
    @staticmethod
    def _mk(severity, type_, title, description, suggestion, date=None, account=None) -> dict:
        """构造异常发现项"""
        return {
            "severity": severity,
            "type": type_,
            "title": title,
            "description": description,
            "suggestion": suggestion,
            "date": date,
            "account": account,
        }

    @staticmethod
    def _group_by_account(records: list) -> dict:
        by_acc = defaultdict(list)
        for r in records:
            acc = r.get("mapped", {}).get("account")
            if acc:
                by_acc[acc].append(r)
        return dict(by_acc)

    def _filter(self, account: Optional[str] = None, days: Optional[int] = None) -> list:
        """按账户与最近 N 天过滤记录（基于 mapped 域）"""
        records = self._load()
        if account:
            records = [r for r in records if r.get("mapped", {}).get("account") == account]
        if days:
            cutoff = (datetime.now() - timedelta(days=days)).date().isoformat()
            records = [r for r in records if r.get("mapped", {}).get("date", "") >= cutoff]
        return records

    # ==================== 示例数据生成 ====================
    def generate_sample(self) -> dict:
        """生成示例投放数据（sample=true 标记，用于演示/测试）
        注入可控异常模式，便于直观验证异常检测：
          - 「信息流-素材测试」某日花费突增 3 倍+
          - 「海外-Meta」末段 CTR 骤降
          - 「搜索-精准」末段连续无转化
        """
        cfg = settings.audit_sample
        accounts = cfg.get("accounts", ["主账户-品牌", "搜索-精准", "信息流-素材测试", "海外-Meta"])
        days = int(cfg.get("days", 30))
        base_imp = float(cfg.get("base_impressions", 80000))
        base_spend = float(cfg.get("base_spend", 800))
        min_conv = int(cfg.get("min_conversions", 4))

        rnd = random.Random(20260805)  # 固定种子，保证可复现
        records = []
        end = date.today()
        for acc_idx, acc in enumerate(accounts):
            acc_imp = base_imp * (0.5 + 0.5 * acc_idx)  # 不同账户规模
            acc_spend = base_spend * (0.6 + 0.4 * acc_idx)
            for i in range(days):
                d = end - timedelta(days=days - 1 - i)
                # 周末/工作日波动 + 随机噪声
                weekday_factor = 1.15 if d.weekday() >= 5 else 1.0
                impressions = int(acc_imp * weekday_factor * rnd.uniform(0.75, 1.25))
                ctr = rnd.uniform(0.8, 2.2) / 100  # 0.8% ~ 2.2%
                clicks = int(impressions * ctr)
                cpc = rnd.uniform(1.2, 3.5)
                spend = round(clicks * cpc, 2)
                # 转化率 2% ~ 8%，保底
                cvr = rnd.uniform(0.02, 0.08)
                conversions = max(min_conv, int(clicks * cvr))
                conversion_value = round(conversions * rnd.uniform(30, 90), 2)

                # ---- 异常注入 ----
                # 1) 信息流-素材测试：第 18-20 天前后花费突增（预算误配，转化未同步）
                if acc == "信息流-素材测试" and 18 <= i <= 20:
                    spend = round(spend * 5.0, 2)  # 确定性 5 倍，确保越过 3 倍阈值
                    conversion_value = round(conversion_value * 0.4, 2)  # 价值没跟上
                # 2) 海外-Meta：最后 4 天 CTR 骤降（素材疲劳）
                if acc == "海外-Meta" and i >= days - 4:
                    clicks = max(2, int(clicks * 0.4))
                    spend = round(clicks * cpc, 2)
                    conversions = max(0, int(conversions * 0.3))
                # 3) 搜索-精准：最后 3 天花费保留但转化归零（转化追踪/落地页问题）
                if acc == "搜索-精准" and i >= days - 3:
                    conversions = 0
                    conversion_value = 0

                records.append({
                    "raw": {
                        "account": acc,
                        "date": d.isoformat(),
                        "impressions": impressions,
                        "clicks": clicks,
                        "conversions": conversions,
                        "spend": spend,
                        "conversion_value": conversion_value,
                    },
                    "mapped": {
                        "account": acc,
                        "date": d.isoformat(),
                        "impressions": impressions,
                        "clicks": clicks,
                        "conversions": conversions,
                        "spend": spend,
                        "conversion_value": conversion_value,
                    },
                    "tags": {},
                    "sample": True,
                })
        with self._lock:
            self._archive_snapshot("sample")  # Phase11: 变更前归档
            self._save(records)
        log_collector.info(EVENT_SYSTEM, f"已生成审计示例数据: {len(records)} 条 / {len(accounts)} 账户 / {days} 天")
        return {"ok": True, "imported": len(records), "sample": True}

    # ==================== 标签库管理（Phase2） ====================
    @property
    def _tag_lib_file(self) -> Path:
        return settings.audit_tag_lib_file

    def _load_tag_lib(self) -> dict:
        """加载用户自定义标签库（覆盖/扩展 spec 预设）"""
        if self._tag_lib_file.exists():
            try:
                return json.loads(self._tag_lib_file.read_text(encoding="utf-8"))
            except (json.JSONDecodeError, OSError):
                pass
        # 初始化：用 spec 预设标签组
        lib = {"groups": settings.audit_tag_groups}
        return lib

    def _save_tag_lib(self, lib: dict):
        self._tag_lib_file.parent.mkdir(parents=True, exist_ok=True)
        self._tag_lib_file.write_text(json.dumps(lib, ensure_ascii=False, indent=2), encoding="utf-8")

    def get_tag_library(self) -> dict:
        """获取标签库（预设 + 用户自定义合并）"""
        return self._load_tag_lib()

    def add_tag_to_group(self, group_id: str, tag: str) -> dict:
        """向标签组添加一个标签值"""
        lib = self._load_tag_lib()
        for g in lib.get("groups", []):
            if g["id"] == group_id:
                if tag not in g.get("tags", []):
                    g.setdefault("tags", []).append(tag)
                    self._save_tag_lib(lib)
                return {"ok": True, "group": g}
        return {"ok": False, "error": f"标签组 {group_id} 不存在"}

    def add_tag_group(self, group_id: str, name: str, description: str = "") -> dict:
        """新增一个标签组"""
        lib = self._load_tag_lib()
        for g in lib.get("groups", []):
            if g["id"] == group_id:
                return {"ok": False, "error": f"标签组 {group_id} 已存在"}
        lib.setdefault("groups", []).append({
            "id": group_id, "name": name, "description": description, "tags": []
        })
        self._save_tag_lib(lib)
        return {"ok": True, "group": lib["groups"][-1]}

    # ==================== 行级打标（Phase2） ====================
    def batch_tag(self, row_indices: list, group_id: str, tags: list, mode: str = "add") -> dict:
        """批量打标：对指定行追加/替换/移除标签
        row_indices: 行索引列表（0-based，对应 _load() 返回的列表索引）
        group_id: 标签组 ID
        tags: 标签值列表
        mode: add(追加) / replace(替换) / remove(移除指定标签) / clear(清空该组)
        """
        with self._lock:
            records = self._load()
            tagged_count = 0
            for idx in row_indices:
                if 0 <= idx < len(records):
                    rec = records[idx]
                    rec.setdefault("tags", {})
                    if mode == "clear":
                        rec["tags"].pop(group_id, None)
                    elif mode == "remove":
                        existing = rec["tags"].get(group_id, [])
                        rec["tags"][group_id] = [t for t in existing if t not in tags]
                        if not rec["tags"][group_id]:
                            rec["tags"].pop(group_id)
                    elif mode == "replace":
                        rec["tags"][group_id] = list(tags)
                    else:  # add
                        existing = rec["tags"].get(group_id, [])
                        for t in tags:
                            if t not in existing:
                                existing.append(t)
                        rec["tags"][group_id] = existing
                    tagged_count += 1
            self._archive_snapshot("tag")  # Phase11: 变更前归档
            self._save(records)
        log_collector.info(EVENT_CONFIG, f"批量打标: {tagged_count} 行 / 组={group_id} / 模式={mode}", {
            "row_count": tagged_count, "group": group_id, "tags": tags, "mode": mode,
        })
        return {"ok": True, "tagged": tagged_count, "group_id": group_id, "mode": mode}

    def get_records_with_tags(self, account: Optional[str] = None, days: Optional[int] = None,
                              limit: int = 500, offset: int = 0) -> dict:
        """获取原始记录（含 raw + tags），分页"""
        records = self._filter(account=account, days=days)
        total = len(records)
        # 分页
        page = records[offset:offset + limit]
        # 返回 raw + tags + 索引（供前端多选用）
        result = []
        for i, r in enumerate(page):
            result.append({
                "index": offset + i,
                "raw": r.get("raw", {}),
                "mapped": r.get("mapped", {}),
                "tags": r.get("tags", {}),
                "sample": r.get("sample", False),
            })
        return {"records": result, "total": total, "limit": limit, "offset": offset}

    def build_report_text(self, account: Optional[str] = None, days: Optional[int] = None) -> str:
        """生成 Markdown 审计报告（供 MCP 业务集成导出，Phase9）
        结构：概览（指标/健康分）→ 异常信号 → 账户对比（如有）
        """
        summary = self.summary(account=account, days=days)
        metrics = summary.get("metrics", {})
        health = summary.get("health", {})
        signals = self.detect_signals()
        day_str = f"近 {days} 天" if days else "全部时间"
        acct_str = f"账户「{account}」" if account else "全部账户"

        def fmt(v):
            if isinstance(v, float):
                return f"{v:.2f}"
            return str(v)

        lines = [
            f"# 广告账户审计报告",
            "",
            f"> {acct_str} · {day_str} · 生成时间 {time.strftime('%Y-%m-%d %H:%M')}",
            "",
            "## 一、数据概览",
            "",
            f"- 覆盖账户：**{metrics.get('account_count', 0)}** 个",
            f"- 覆盖天数：**{metrics.get('day_count', 0)}** 天",
            f"- 数据行数：**{metrics.get('record_count', 0)}** 条",
        ]
        for k, label in (("spend", "总花费"), ("impressions", "总曝光"),
                         ("clicks", "总点击"), ("conversions", "总转化")):
            if k in metrics:
                lines.append(f"- {label}：**{fmt(metrics[k])}**")
        for k, label in (("ctr", "平均 CTR"), ("cvr", "平均 CVR"), ("cpa", "平均 CPA"), ("roas", "平均 ROAS")):
            if k in metrics:
                lines.append(f"- {label}：**{fmt(metrics[k])}**")

        lines += ["", "## 二、健康评分", ""]
        if health:
            lines.append(f"- 健康分：**{health.get('score', '-')}/100**（{health.get('grade', '')}）")
            for k, v in (health.get("breakdown") or {}).items():
                lines.append(f"  - {k}：{v}")
        lines.append(f"- 异常信号：**{summary.get('anomaly_count', 0)}** 条"
                     f"（高 {summary.get('anomaly_by_severity', {}).get('high', 0)} / "
                     f"中 {summary.get('anomaly_by_severity', {}).get('medium', 0)} / "
                     f"低 {summary.get('anomaly_by_severity', {}).get('low', 0)}）")

        medium_high = [s for s in signals if s.get("severity") in (SEV_CRITICAL, SEV_HIGH, SEV_MEDIUM)]
        lines += ["", "## 三、异常信号", ""]
        if not medium_high:
            lines.append("无中/高风险信号，投放健康。")
        for s in medium_high[:10]:
            lines.append(f"- **[{s.get('severity', '').upper()}]** {s.get('account', '') or ''} · {s.get('category', '')}")
            lines.append(f"  - {s.get('description', '')}")
            if s.get("impact"):
                lines.append(f"  - 影响：{s.get('impact')}")
            if s.get("suggestion"):
                lines.append(f"  - 建议：{s.get('suggestion')}")

        lines += ["", "---", "由 AdToEarn WebUI 自动生成"]
        return "\n".join(lines)


# 全局单例
audit_service = AuditService()
